"""********************************************
File: client.py
Author: Mahdi Habibi

Desc. :
    Obtains merged data by making a POST call 
    to server.py containing a csv file. 
    Then applies filters to the data and output
    as specified in the task.
********************************************"""

# Built-ins
from io import BytesIO
from datetime import date # for today
import argparse

# Third-party
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


GREEN = "007500"
ORANGE = "FFA500"
RED = "b30000"

def df_towb(df):
    """
    Writes a pandas
    dataframe df into a
    openpyxl's workbook
    in memory.
    """

    xlbuf = BytesIO() # create buffer
    df.to_excel(xlbuf, index=False) # out xlsx
    xlbuf.seek(0) # seek begin
    wb = load_workbook(xlbuf) # load wb.
    return wb

def tint_labelids(ws):
    """ 
    Given a openpyxl.workbook wb
    Paint the labelId cell to the
    color assiciated with it and
    return the new painted workbook.
    """

    i = 2 # skip the keys
    j = 0
    while i < ws.max_row and\
            j < ws.max_column:
        if ws[1][j].value != "labelIds":
            j += 1
            continue

        color = ws[i][j].value
        if color is not None:
            # delete hashtag
            color = str(color)[1:]

            fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid")
            ws[i][j].fill = fill

        i += 1

def hu_to_color(entry):
    """
    Compare dates and return
    the proper color, regarding
    speciefications of task.

    P.N:
    We could've use any api to
    compare date types, but it
    doesn't hurt to write a
    little algorithm!
    """

    today = int(str(
        date.today()).replace(
            "-", ""))
    twelve_m = today - 10000
    three_m = today - 300

    # parse hu date
    hu = entry.value
    hu = int(hu.replace("-", ""))

    if hu > twelve_m:
        if hu > three_m:
            return GREEN
        return ORANGE
    return RED

def paint_row(row, color):
    """
    Paints all cells of a row.
    """

    # Obtain fill
    fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid")

    # Apply
    for cell in row:
        cell.fill = fill

def paint_rows(ws):
    """
    Use the hu values to color
    each row, as speciefied in
    the task.
    """

    # Index of 'hu' column
    hu_i = 0
    while hu_i < ws.max_column and\
    ws[1][hu_i].value != "hu":
        hu_i += 1

    # Iterate, evaluate and paint
    for row in ws.iter_rows(min_row=2):
        entry = row[hu_i]
        paint_row(row, 
                  hu_to_color(entry))

def client():
    """
    Transmits a CSV to
    a REST-API (s. Server-section below),
    handles the response and
    generates an Excel-File
    taking the input parameters
    into account.
    """

    # Handle args
    parser = argparse.ArgumentParser(
            description=\
                    "client.py(VERO pytask):"
                    " write vehicles data"
                    " to a xlsx file.")

    # Add first
    parser.add_argument("filename")

    # kolumns always contain rnr field
    cols = ["rnr"]
    # Handle -k flag
    parser.add_argument(
            "-k",
            "--keys",
            nargs="+",
            help="Columns to write"
            " out : arbitary amount"
            " of strings")

    parser.add_argument(
            "-c",
            "--colored",
            action="store_true",
            help="Disable colord output.",
            default=False)

    args = parser.parse_args()

    # Call server and retrieve json data
    response = requests.post(
      url="http://127.0.0.1:5000/vehicles",
      files={"file": open(args.filename, "rb")})
    df = pd.read_json(response.json())

    # Additional columns
    for arg in args.keys:
        # Check input in O(n)
        if arg not in df.columns:
            print(f"no column matching the"
                  f"{arg}, will leave out")
            continue
        cols.append(arg)

    # Sort by gruppe
    df = df.sort_values(by=["gruppe"])

    # Obtain df with additional columns
    df = df[[col for col in cols]]
    del cols

    # Get openpyxl Workbook for colors
    wb = df_towb(df)
    del df
    ws = wb["Sheet1"]

    # Tint cell with the resolved
    # label id value
    tint_labelids(ws)

    # Paint each row for colored
    if (args.colored):
        paint_rows(ws)

    # Output
    wb.save(f"vehicles_"
            f"{str(date.today())}"
            ".xlsx")

client()
