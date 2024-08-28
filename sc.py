"""********************************************
File: server-client.py
Author: Metflekx

Desc. :
    Write two python scripts that
    have to achieve the common goal
    to downloads a certain set of resources
    , merges them with CSV-transmitted
    resources, and converts them to a
    formatted excel file. In particular,
    the script should:
********************************************"""

# Built-ins
import sys # For argv
from io import BytesIO
from datetime import date # for today

# Third-party
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def get_access_token():
    """Get access token for
    the API
    """

    url = "https://api.baubuddy.de/index.php/login"
    payload = {
        "username": "365",
        "password": "1"
    }
    headers = {
        "Authorization":
          "Basic QVBJX0V4cGxvcmVyOjEyMzQ1NmlzQUxhbWVQYXNz",
        "Content-Type":
          "application/json"
    }
    res = requests.request(
      "POST", url,
      json=payload,
      headers=headers)
    return res.json()['oauth']['access_token']

def server(csvpath):
    """ This script should offer
    a REST-API, that accepts a CSV,
    downloads a certain set of
    resources, merges them with the
    CSV, applies filtering, and
    returns them in an appropriate
    data-structure

    Takes path to a csv: reqpath,
    makes a reqdf data frame that,
    gets and reads resources into
    apidf. merges into resdf, applies
    filtering and returns resdf in json.
    """

    # TODO:
    # [ ] make sure the result is distinct
    # [ ] REST-API (e.g. FastAPI, Flask, Django …)
            # offering a POST Call which accepts a
            # transmitted CSV containing vehicle information

    # Read the data sent over request
    reqdf = pd.read_csv(csvpath, delimiter=';')
    # Request the resources
    acctoken = get_access_token()

    res = (requests.get(
            "https://api.baubuddy.de"
            "/dev/index.php/v1/vehicles"
            "/select/active",
            headers = {
                "Authorization":
                f"Bearer {acctoken}"}))
    apidf = pd.DataFrame(res.json())

    # Filter out any resources
    # that do not have a value
    # set for hu field
    apidf = apidf[apidf["hu"].notna()]

    # merge req and res data
    resdf = pd.merge(reqdf, apidf,
                     on="kurzname",
                     how="outer",
                     validate="one_to_many",
                     suffixes=("", "_res"))

    # Combine labelId and LabelId_res
    resdf['labelIds'] = resdf[
            'labelIds'].combine_first(
                    resdf['labelIds_res'])
    resdf = resdf.drop(
            columns=['labelIds_res'])

    # For each labelId in the
    # vehicle's JSON array
    # labelIds resolve its colorCode
    for i, row in resdf.iterrows():
        # Skip NaN
        if pd.notna(row["labelIds"]):
            res = requests.get(
                "https://api.baubuddy.de/"
                "dev/index.php/v1/labels/"
                f"{row['labelIds']}",
            headers = {
              "Authorization":
                f"Bearer {acctoken}"})

            # bad response
            if res.status_code != 200:
                continue

            # set colorCode
            resdf.at[i, "labelIds"]=\
            res.json()[0]["colorCode"]

    # return data-structure in JSON format
    return resdf.to_json()


GREEN = "007500"
ORANGE = "FFA500"
RED = "b30000"


def df_towb(df):
    """ writes a pandas
    dataframe df into a
    openpyxl's workbook
    in memory.
    """

    xlbuf = BytesIO() # create buffer
    df.to_excel(xlbuf, index=False) # out xlsx
    xlbuf.seek(0) # seek begin
    wb = load_workbook(xlbuf) # load wb.
    return wb

def get_colors(df):
    """ Make a new column colortmp
    in df Obtain the row color 
    according to the hu value.
    """

    df["colortmp"] = None 
    today = int(str(
        date.today()).replace(
            "-", ""))
    twelve_m = today - 10000
    three_m = today - 300
    for i, huval in enumerate(df["hu"]):
        if isinstance(huval, str):
            huval = int(
                    huval.replace(
                        "-", "")) 
            if huval < twelve_m:
                if huval < three_m:
                    df.at[i, "colortmp"] = GREEN
                    continue
                df.at[i, "colortmp"] = ORANGE
                continue 
            df.at[i, "colortmp"] = RED

def tint_labelids(ws):
    """ If labelIds are given 
    and at least one colorCode 
    could be resolved, use the 
    first colorCode to tint the 
    cell's text 
    (if labelIds is given in -k)

    Given a openpyxl.workbook wb
    Paint the labelId cell to the
    color assiciated with it and 
    return the new painted workbook.
    """

    i = 2 # skip the keys
    j = 0
    while i < ws.max_row and\
            j < ws.max_column + 1:
        if ws[1][j].value != "labelIds":
            j += 1
            continue

        color = ws[i][j].value
        if color is not None:
            # delete hashtag
            color = str(color)[1:]

            fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid")
            ws[i][j].fill = fill
        
        i += 1

def paint_rows(ws) -> int:
    """ Use the values stored
    in colortmp, to paint each
    row according to hu value.
    returns the index at which
    the colortmp exist.
    in a call to paint_rows the
    return value must be used
    to delete tmp column.
    """

    # find the colortmp index
    cti = 0
    while cti < ws.max_column and\
    ws[1][cti].value != "colortmp":
        cti += 1

    i = 2
    while i < ws.max_row + 1:
        # Get color
        color = ws[i][cti].value
        if isinstance(color, str):
            # Paint row
            j = 0
            fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid")
            while j < ws.max_column:
                ws[i][j].fill = fill
                j += 1

        i += 1

    # return index of tmp col
    return cti + 1

def client():
    """Transmits a CSV to
    a REST-API
    (s. Server-section below),
    handles the response and
    generates an Excel-File
    taking the input parameters
    into account.

    TODO:
        [*] color cells
        [ ] handle args
        [ ] seperate files
        [ ] after server listening
            make calls
    """

    # Call server and get the merged
    # data in json
    df = pd.read_json(
            server("vehicles.csv"))

    # handle -k flag
    # Columns always contain rnr field
    cols = ["rnr"]
    # Only keys that match the input
    # arguments are considered as
    # additional columns
    for arg in sys.argv[1:]:
        # Check for valid input:
        # O(N) lookup OK since
        # n is samll for columns
        if arg not in df.columns:
            print(f"no column matching the"
                  f"{arg}, will leave out")
            continue
        cols.append(arg)

    # -c flag
    get_colors(df)
    cols.append("colortmp")

    # Rows are sorted by response
    # field gruppe
    df = df.sort_values(by=["gruppe"])

    df = df[[col for col in cols]]
    del cols # TODO: later after restapi implementation handle the scope

    wb = df_towb(df)
    del df

    # Get worksheet
    ws = wb["Sheet1"]

    # Tint cell with the resolved
    # label id value
    tint_labelids(ws)

    # Paint each row for colored
    # output
    ws.delete_cols(
            paint_rows(ws), amount=1)
    sys.exit(0)

    # [ ] Take an input parameter -k/--keys that can receive an arbitrary amount of string arguments
    # [ ] Take an input parameter -c/--colored that receives a boolean flag and defaults to True
    # [ ] Transmit CSV containing vehicle information to the POST Call of the server (example data: vehicles.csv)
    # [ ] Convert the servers response into an excel file that contains all resources and make sure that:
    # [ ] The file should be named vehicles_{current_date_iso_formatted}.xlsx

client()
